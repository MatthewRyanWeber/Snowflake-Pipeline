"""Excel (.xlsx) source — a worksheet as a stand-in relational table. openpyxl imported lazily.

The 'table' names the worksheet (falls back to the active sheet); the first row is the header.
Rows are filtered by the high-water-mark and ordered ascending, mirroring the SQL sources, so
incremental checkpointing stays correct. Same fetch_batches / count contract as the DB sources.

Bounded-memory by design: a worksheet is read whole and sorted in memory (openpyxl gives no
sorted server-side cursor), so this suits reference/export files, not billion-row streams — the
DB sources cover those with server-side cursors.
"""

import logging
from pathlib import Path

from .ordering import hwm_gt, hwm_key

logger = logging.getLogger(__name__)


class ExcelSource:
    def __init__(self, path: str):
        self.path = Path(path)

    def connect(self):
        if not self.path.exists():
            raise FileNotFoundError(f"excel source not found: {self.path}")
        logger.info("excel source: %s", self.path)
        return self

    def _rows(self, table):
        import openpyxl

        wb = openpyxl.load_workbook(self.path, read_only=True, data_only=True)
        try:
            ws = wb[table] if table in wb.sheetnames else wb.active
            it = ws.iter_rows(values_only=True)
            header = next(it, None)
            if header is None:
                return []
            header = [str(h) for h in header]
            return [dict(zip(header, r)) for r in it]
        finally:
            wb.close()

    def fetch_batches(self, table: str, hwm_column: str, since, batch_size: int):
        rows = self._rows(table)
        if not rows:
            return
        if hwm_column not in rows[0]:
            raise ValueError(f"hwm_column {hwm_column!r} not in sheet {table!r} columns")
        rows.sort(key=lambda r: hwm_key(r[hwm_column]))
        if since is not None:
            rows = [r for r in rows if hwm_gt(r[hwm_column], since)]
        for i in range(0, len(rows), batch_size):
            yield rows[i:i + batch_size]

    def count(self, table: str, hwm_column: str, since) -> int:
        rows = self._rows(table)
        if not rows:
            return 0
        if since is None:
            return len(rows)
        return sum(1 for r in rows if hwm_gt(r[hwm_column], since))

    def close(self):
        pass
